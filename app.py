import pandas as pd
from openpyxl.styles import Font
from openpyxl import load_workbook
import streamlit as st
import io

st.title("📊 Jira CSV to Excel Converter")

uploaded_file = st.file_uploader("Upload your Jira CSV file", type="csv")

if uploaded_file:
    df = pd.read_csv(uploaded_file)

    fields = [
        'Issue key', 
        'Summary',
        'Custom field (Sub-Project)',
        'Issue Type', 
        'Status',
        'Created',
        'Updated',
        'Priority',
        'Fix versions'
    ]

    df_selected = df[fields]

    df_selected.columns = [
        'Key', 
        'Summary',
        'Sub-Project',
        'Issue Type', 
        'Status', 
        'Updated',
        'Created',
        'Priority',
        'Fix Version/s'
    ]

    jira_base_url = 'https://ngsoft.atlassian.net/browse/'
    df_selected['Link'] = jira_base_url + df_selected['Key']

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_selected.to_excel(writer, index=False)
        ws = writer.book.active
        link_column = ws.max_column

        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=link_column)
            issue_key = ws.cell(row=row, column=1).value
            hyperlink = f'{jira_base_url}{issue_key}'
            cell.value = issue_key
            cell.hyperlink = hyperlink
            cell.font = Font(color='0000FF', underline='single')

    st.success("✅ File processed!")
    st.download_button(
        label="📥 Download Excel",
        data=output.getvalue(),
        file_name="output_custom_jira_issues.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
